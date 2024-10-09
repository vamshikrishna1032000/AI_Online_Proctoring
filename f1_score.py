import cv2
import numpy as np
import math
from eye_tracker import contouring, eye_on_mask, print_eye_pos, process_thresh
from face_detector import get_face_detector, find_faces
from face_landmarks import draw_marks, get_landmark_model, detect_marks
from head_pose_estimation import head_pose_points
import time
import csv
import re
from openpyxl import load_workbook
# from sklearn.metrics import f1_score


ResultBook=load_workbook('result.xlsx')
ResultSheet=ResultBook.active



y_true = []
y_pred = []
for col in ResultSheet.iter_cols(2, 2):
    for row in range(1, ResultSheet.max_row):
        if(col[row].value=='cheating'):
            y_true.append(1)
        elif(col[row].value=='not cheating'):
            y_true.append(0)  
        else:
            print(row,col[row].value)
for col in ResultSheet.iter_cols(3, 3):
    for row in range(1, ResultSheet.max_row):
        if(col[row].value=='cheating'):
            y_pred.append(1)
        elif(col[row].value=='not cheating') :
            y_pred.append(0)  
        else:
            print(row,col[row].value)

print(len(y_true),len(y_pred))


